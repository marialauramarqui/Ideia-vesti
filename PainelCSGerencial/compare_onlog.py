"""
Conferencia da planilha do Diogo (Onlog) vs dados do Fabric (onlog_data.json).

Uso:
    py compare_onlog.py "C:\\caminho\\Fechamento ONLOG.xlsx"
    py compare_onlog.py "...xlsx" --de 2026-04-01 --ate 2026-04-15

Sem --de/--ate, detecta a quinzena automaticamente pela menor/maior data da planilha.

Saida:
    - relatorio_onlog.html  (abra no navegador - mesmo conteudo do dashboard)
    - divergencias.csv, so_planilha.csv, so_fabric.csv
"""

import argparse
import csv
import json
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl nao instalado. Rode: py -m pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).parent
ONLOG_JSON = ROOT / "onlog_data.json"


def norm_txt(s) -> str:
    if s is None:
        return ""
    s = str(s)
    nfd = unicodedata.normalize("NFD", s)
    no_diac = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    out = []
    for c in no_diac.upper():
        if c.isalnum():
            out.append(c)
        else:
            out.append(" ")
    return " ".join("".join(out).split())


def norm_uf(s) -> str:
    return (str(s or "").strip().upper())[:2]


def parse_val_br(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def fmt_brl(v) -> str:
    if v is None:
        return "-"
    return "R$ " + f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def read_planilha(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        rec = {header[i]: r[i] for i in range(len(header))}
        out.append(rec)
    return out


def aggregate_planilha(rows: list[dict], de: str, ate: str) -> dict:
    by = {}
    for r in rows:
        cv = str(r.get("CodigoVolume") or "").strip()
        if not cv or "_" not in cv:
            continue
        d = r.get("Data")
        d_str = ""
        if isinstance(d, datetime):
            d_str = d.date().isoformat()
        elif isinstance(d, str):
            d_str = d[:10]
        if de and d_str and d_str < de:
            continue
        if ate and d_str and d_str > ate:
            continue
        if cv not in by:
            dom, order = cv.split("_", 1)
            by[cv] = {
                "codigoVolume": cv,
                "orderNumber": r.get("NumeroPedido") or order,
                "domainId": dom,
                "cliente": r.get("Destinatario") or "",
                "cidade": r.get("CidadeDestinatario") or "",
                "uf": r.get("UFDestinatario") or "",
                "status": r.get("Status") or "",
                "data": d_str,
                "postagem": 0.0,
                "_n": 0,
            }
        v = parse_val_br(r.get("ValorPostagem"))
        if v is not None:
            by[cv]["postagem"] += v
        by[cv]["_n"] += 1
    return by


def filter_fabric(pedidos: list[dict], de: str, ate: str) -> dict:
    out = {}
    for p in pedidos:
        d = p.get("data") or ""
        if not d or d < de or d > ate:
            continue
        k = f'{p.get("dominioId","")}_{p.get("orderNumber","")}'
        out[k] = p
    return out


def detect_quinzena(rows: list[dict]) -> tuple[str, str]:
    min_d, max_d = None, None
    for r in rows:
        d = r.get("Data")
        s = ""
        if isinstance(d, datetime):
            s = d.date().isoformat()
        elif isinstance(d, str):
            s = d[:10]
        if not s:
            continue
        if min_d is None or s < min_d:
            min_d = s
        if max_d is None or s > max_d:
            max_d = s
    if not min_d:
        return "", ""
    y, mo, dia = int(min_d[:4]), int(min_d[5:7]), int(min_d[8:10])
    mes = min_d[:7]
    if dia <= 15:
        return f"{mes}-01", f"{mes}-15"
    # 2a: ate ultimo dia do mes
    if mo == 12:
        last = 31
    else:
        from calendar import monthrange
        last = monthrange(y, mo)[1]
    return f"{mes}-16", f"{mes}-{last:02d}"


def compare(planilha: dict, fabric: dict):
    keys = set(planilha) | set(fabric)
    dif = []
    only_p = []
    only_f = []
    ok = 0
    for k in keys:
        pl = planilha.get(k)
        fa = fabric.get(k)
        if pl and not fa:
            only_p.append(pl)
            continue
        if fa and not pl:
            only_f.append(fa)
            continue
        divs = []
        if norm_txt(pl["cliente"]) != norm_txt(fa.get("cliente", "")):
            divs.append(("Cliente", pl["cliente"], fa.get("cliente", "")))
        dest_p = f'{norm_txt(pl["cidade"])} / {norm_uf(pl["uf"])}'
        dest_f = f'{norm_txt(fa.get("cidade",""))} / {norm_uf(fa.get("uf",""))}'
        if dest_p != dest_f:
            divs.append((
                "Destino",
                f'{pl["cidade"]}/{pl["uf"]}',
                f'{fa.get("cidade","")}/{fa.get("uf","")}',
            ))
        if fa.get("cancelado"):
            divs.append(("Cancelado", "(postado pela Onlog)", "CANCELADO no Vesti"))
        pp = pl["postagem"]
        pf = fa.get("valorPostagem")
        if pp is not None and pf is not None:
            if abs(pp - pf) > 0.01:
                divs.append(("Valor Postagem", fmt_brl(pp), fmt_brl(pf)))
        elif pp is not None and pf is None:
            divs.append(("Valor Postagem", fmt_brl(pp), "(sem dado no Fabric)"))
        if divs:
            for campo, a, b in divs:
                dif.append({
                    "orderNumber": fa.get("orderNumber"),
                    "marca": fa.get("marca", "-"),
                    "campo": campo,
                    "planilha": a,
                    "fabric": b,
                })
        else:
            ok += 1
    return ok, dif, only_p, only_f


HTML_TPL = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Conferencia Onlog {de} a {ate}</title>
<style>
body{{font-family:Inter,Arial,sans-serif;background:#F8F9FA;margin:0;padding:20px;color:#2D3436}}
h1{{margin:0 0 4px}} h2{{margin:24px 0 8px;font-size:18px}}
.kpi{{display:inline-block;background:#FFF;border-radius:12px;padding:14px 20px;margin:6px 8px 0 0;
     box-shadow:0 2px 8px rgba(0,0,0,.06);min-width:160px}}
.kpi .l{{font-size:11px;color:#636E72;text-transform:uppercase;letter-spacing:.5px}}
.kpi .v{{font-size:28px;font-weight:800;margin-top:4px}}
.kpi .s{{font-size:11px;color:#636E72}}
.green{{color:#00B894}} .red{{color:#E17055}} .orange{{color:#F39C12}} .blue{{color:#0984E3}}
table{{border-collapse:collapse;width:100%;background:#FFF;border-radius:8px;overflow:hidden;
       box-shadow:0 2px 8px rgba(0,0,0,.05);margin-top:6px;font-size:12px}}
th{{background:#F0EFFF;text-align:left;padding:10px 12px;font-weight:600}}
td{{padding:8px 12px;border-top:1px solid #DFE6E9}}
.right{{text-align:right}} .mono{{font-family:Consolas,monospace}}
.summary{{color:#636E72;margin-top:8px}}
details{{background:#FFF;border-radius:8px;padding:10px 14px;margin-top:10px;
         box-shadow:0 2px 8px rgba(0,0,0,.05)}}
summary{{cursor:pointer;font-weight:700;font-size:14px}}
</style></head><body>
<h1>Conferencia ONLOG &middot; {de} a {ate}</h1>
<p class="summary">{n_planilha} pedido(s) na planilha &middot; {n_fabric} pedido(s) no Fabric &middot; gerado em {now}</p>

<div class="kpi"><div class="l">Pedidos batendo</div><div class="v green">{ok}</div><div class="s">cliente, destino, cancelado e postagem conferem</div></div>
<div class="kpi"><div class="l">Divergencias</div><div class="v red">{n_dif_uniq}</div><div class="s">{n_dif} apontamento(s) em {n_dif_uniq} pedido(s)</div></div>
<div class="kpi"><div class="l">So na planilha</div><div class="v orange">{n_only_p}</div><div class="s">nao achei no Fabric (Mongo)</div></div>
<div class="kpi"><div class="l">So no Fabric</div><div class="v blue">{n_only_f}</div><div class="s">Onlog nao postou ainda / cancelado</div></div>

<details open><summary class="red">Divergencias por campo - {n_dif}</summary>
<table><thead><tr><th>Pedido</th><th>Marca</th><th>Campo</th><th>Planilha (Diogo)</th><th>Fabric (Vesti)</th></tr></thead>
<tbody>{rows_dif}</tbody></table></details>

<details><summary class="orange">Pedidos so na planilha - {n_only_p}</summary>
<table><thead><tr><th>CodigoVolume</th><th>Pedido</th><th>Cliente</th><th>Destino</th><th>Status</th><th class="right">Postagem</th></tr></thead>
<tbody>{rows_only_p}</tbody></table></details>

<details><summary class="blue">Pedidos so no Fabric - {n_only_f}</summary>
<table><thead><tr><th>Data</th><th>Pedido</th><th>Marca</th><th>Cliente</th><th>Destino</th><th>Status</th><th class="right">Valor</th></tr></thead>
<tbody>{rows_only_f}</tbody></table></details>

</body></html>"""


def esc(s) -> str:
    if s is None:
        return ""
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def render_html(de, ate, ok, dif, only_p, only_f, n_planilha, n_fabric) -> str:
    rows_dif = "".join(
        f'<tr><td class="mono">#{esc(d["orderNumber"])}</td>'
        f'<td><b>{esc(d["marca"])}</b></td>'
        f'<td class="red"><b>{esc(d["campo"])}</b></td>'
        f'<td style="color:#636E72">{esc(d["planilha"])}</td>'
        f'<td>{esc(d["fabric"])}</td></tr>'
        for d in dif
    ) or '<tr><td colspan="5" style="text-align:center;color:#999;padding:14px">Nenhuma divergencia.</td></tr>'

    rows_only_p = "".join(
        f'<tr><td class="mono">{esc(p["codigoVolume"])}</td>'
        f'<td class="mono">#{esc(p["orderNumber"])}</td>'
        f'<td>{esc(p["cliente"])}</td>'
        f'<td>{esc(p["cidade"])}/{esc(p["uf"])}</td>'
        f'<td style="color:#636E72;font-size:11px">{esc(p["status"])}</td>'
        f'<td class="right">{esc(fmt_brl(p["postagem"]))}</td></tr>'
        for p in only_p
    ) or '<tr><td colspan="6" style="text-align:center;color:#999;padding:14px">Nenhum.</td></tr>'

    def status_lbl(p):
        if p.get("cancelado"):
            return "Cancelado"
        return p.get("status") or "-"

    rows_only_f = "".join(
        f'<tr><td>{esc(p.get("data",""))}</td>'
        f'<td class="mono">#{esc(p.get("orderNumber"))}</td>'
        f'<td><b>{esc(p.get("marca","-"))}</b></td>'
        f'<td>{esc(p.get("cliente","-"))}</td>'
        f'<td>{esc(p.get("cidade",""))}/{esc(p.get("uf",""))}</td>'
        f'<td style="color:{"#E17055" if p.get("cancelado") else "#636E72"}">{esc(status_lbl(p))}</td>'
        f'<td class="right">{esc(fmt_brl(p.get("valor")))}</td></tr>'
        for p in only_f
    ) or '<tr><td colspan="7" style="text-align:center;color:#999;padding:14px">Nenhum.</td></tr>'

    n_dif_uniq = len({d["orderNumber"] for d in dif})

    return HTML_TPL.format(
        de=de, ate=ate, ok=ok,
        n_dif=len(dif), n_dif_uniq=n_dif_uniq,
        n_only_p=len(only_p), n_only_f=len(only_f),
        n_planilha=n_planilha, n_fabric=n_fabric,
        rows_dif=rows_dif, rows_only_p=rows_only_p, rows_only_f=rows_only_f,
        now=datetime.now().strftime("%d/%m/%Y %H:%M"),
    )


def write_csv(path: Path, header: list[str], rows: list[list]):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(header)
        w.writerows(rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Caminho da planilha do Diogo (Fechamento ONLOG.xlsx)")
    ap.add_argument("--de", help="Data inicial (YYYY-MM-DD). Default: detectado da planilha.")
    ap.add_argument("--ate", help="Data final (YYYY-MM-DD). Default: detectado da planilha.")
    ap.add_argument("--out", default=str(ROOT / "relatorio_onlog.html"),
                    help="Caminho do HTML de saida.")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERRO: nao achei {xlsx_path}")
        sys.exit(1)
    if not ONLOG_JSON.exists():
        print(f"ERRO: {ONLOG_JSON} nao existe. Rode py fetch_onlog.py antes.")
        sys.exit(1)

    print(f"[1/4] Lendo planilha: {xlsx_path.name}")
    raw = read_planilha(xlsx_path)
    print(f"      {len(raw)} linhas brutas")

    if args.de and args.ate:
        de, ate = args.de, args.ate
    else:
        de, ate = detect_quinzena(raw)
        if not de:
            print("ERRO: nao consegui detectar quinzena pelas datas. Use --de e --ate.")
            sys.exit(1)
        print(f"      quinzena detectada: {de} a {ate}")

    print(f"[2/4] Agregando planilha por CodigoVolume na quinzena")
    planilha = aggregate_planilha(raw, de, ate)
    print(f"      {len(planilha)} pedidos unicos na planilha")

    print(f"[3/4] Lendo Fabric (onlog_data.json)")
    onlog_data = json.loads(ONLOG_JSON.read_text(encoding="utf-8"))
    fabric = filter_fabric(onlog_data["pedidos"], de, ate)
    print(f"      {len(fabric)} pedidos do Fabric na quinzena")

    print(f"[4/4] Comparando")
    ok, dif, only_p, only_f = compare(planilha, fabric)
    print(f"      OK={ok}  Divergencias={len(dif)} (em {len({d['orderNumber'] for d in dif})} pedidos)")
    print(f"      So na planilha={len(only_p)}  So no Fabric={len(only_f)}")

    out_path = Path(args.out)
    out_path.write_text(
        render_html(de, ate, ok, dif, only_p, only_f, len(planilha), len(fabric)),
        encoding="utf-8",
    )
    print(f"\n>> HTML: {out_path}")

    write_csv(ROOT / "divergencias.csv",
              ["Pedido", "Marca", "Campo", "Planilha", "Fabric"],
              [[d["orderNumber"], d["marca"], d["campo"], d["planilha"], d["fabric"]] for d in dif])
    write_csv(ROOT / "so_planilha.csv",
              ["CodigoVolume", "Pedido", "Cliente", "Cidade", "UF", "Status", "Postagem"],
              [[p["codigoVolume"], p["orderNumber"], p["cliente"], p["cidade"], p["uf"], p["status"], p["postagem"]] for p in only_p])
    write_csv(ROOT / "so_fabric.csv",
              ["Data", "Pedido", "Marca", "Cliente", "Cidade", "UF", "Cancelado", "Status", "Valor"],
              [[p.get("data"), p.get("orderNumber"), p.get("marca"), p.get("cliente"),
                p.get("cidade"), p.get("uf"), "SIM" if p.get("cancelado") else "", p.get("status"), p.get("valor")]
               for p in only_f])
    print(f">> CSVs: divergencias.csv, so_planilha.csv, so_fabric.csv (mesma pasta)")


if __name__ == "__main__":
    main()

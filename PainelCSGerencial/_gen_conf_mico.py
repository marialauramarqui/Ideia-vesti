"""Gera planilha Excel de auditoria Mi&Co.

Conferencia da planilha do Diogo (versao Mi&Co Corrigido) vs dados do Mongo (Fabric):
- ValorPostagem (planilha)  = custo bruto Onlog -> Vesti (sem 10%)
- BIA (Mongo)               = o que o cliente pagou no checkout
- Esperado a cobrar Mi&Co   = MAX(BIA, ValorPostagem * 1.10) - garante minimo de 10%
- Margem Vesti              = (a cobrar) - ValorPostagem

Auditoria por pedido:
1. Status atrasado: Mongo em WAITING/SEPARATED mas Onlog ja postou
2. Sem etiqueta no Mongo: postado pelo Diogo mas sem URL etiqueta no Mongo
3. Frete divergente: BIA difere de Postagem*1.10 acima de 5%
"""
import json
from datetime import datetime as _dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX_IN = r'C:/Users/Laura/Downloads/MI E CO Corrigido.xlsx'
ROOT = r'C:/Users/Laura/Projetos/Ideia-vesti/PainelCSGerencial'

src = openpyxl.load_workbook(XLSX_IN, data_only=True)
ws_src = src[src.sheetnames[0]]
header = [c.value for c in ws_src[1]]
def col(c): return header.index(c)

cvs = {}
for row in ws_src.iter_rows(min_row=2, values_only=True):
    cv = str(row[col('CodigoVolume')] or '').strip()
    if not cv or '_' not in cv:
        continue
    cvs.setdefault(cv, {
        'data': row[col('Data')],
        'op': row[col('Operador')],
        'modal': row[col('Modalidade')],
        'cliente': row[col('Destinatario')],
        'cidade': row[col('CidadeDestinatario')],
        'uf': row[col('UFDestinatario')],
        'statusOnlog': row[col('Status')],
        'codigoInterno': row[col('CodigoInterno')],
        'numeroNF': row[col('NumeroNF')],
        'postagem': 0,
    })
    cvs[cv]['postagem'] += float(row[col('ValorPostagem')] or 0)

# Cruza com Mongo
with open(f'{ROOT}/onlog_data.json', 'r', encoding='utf-8') as f:
    od = json.load(f)
mongo_by_key = {f'{p.get("dominioId")}_{p.get("orderNumber")}': p for p in od.get('pedidos', [])}
for cv, c in cvs.items():
    p = mongo_by_key.get(cv) or {}
    c['mongoBia'] = float(p.get('cotacaoBia') or 0)
    c['mongoValor'] = float(p.get('valor') or 0)
    c['mongoStatus'] = p.get('status') or ''
    c['mongoCancelado'] = bool(p.get('cancelado'))
    c['mongoTemEtiqueta'] = bool(p.get('comEtiqueta'))
    c['mongoEtiquetaUrl'] = p.get('etiquetaUrl') or ''
    c['provider'] = (p.get('provider') or '').replace('Vesti - ', '')

# Datas para cabecalho
datas = [c['data'] for c in cvs.values() if c['data']]
de_str = min(datas).strftime('%d/%m/%Y') if datas else '-'
ate_str = max(datas).strftime('%d/%m/%Y') if datas else '-'

# Totais
total_postagem = sum(c['postagem'] for c in cvs.values())
total_bia = sum(c['mongoBia'] for c in cvs.values())
total_minimo = sum(c['postagem'] * 1.10 for c in cvs.values())
total_cobrar = sum(max(c['mongoBia'], c['postagem'] * 1.10) for c in cvs.values())
margem_vesti = total_cobrar - total_postagem

# Auditoria
THRESHOLD_FRETE = 0.05
audit = {'status_atrasado': [], 'sem_etiqueta': [], 'frete_divergente': []}
for cv, c in cvs.items():
    n = cv.split('_')[1]
    # 1) Status atrasado: Mongo WAITING/SEPARATED mas planilha ja postou
    if (c['mongoStatus'] or '').upper() in ('WAITING', 'SEPARATED') and c['statusOnlog']:
        audit['status_atrasado'].append((n, c))
    # 2) Sem etiqueta Mongo
    if not c['mongoTemEtiqueta']:
        audit['sem_etiqueta'].append((n, c))
    # 3) Frete divergente
    if c['mongoBia'] > 0 and c['postagem'] > 0:
        esperado = c['postagem'] * 1.10
        diff_pct = abs(c['mongoBia'] - esperado) / esperado
        if diff_pct > THRESHOLD_FRETE and abs(c['mongoBia'] - esperado) > 1.0:
            audit['frete_divergente'].append((n, c, c['mongoBia'] - esperado, diff_pct))

n_status_atr = len(audit['status_atrasado'])
n_sem_etiq = len(audit['sem_etiqueta'])
n_frete_div = len(audit['frete_divergente'])
pedidos_problema = set()
for n, _ in audit['status_atrasado']: pedidos_problema.add(n)
for n, _ in audit['sem_etiqueta']: pedidos_problema.add(n)
for n, _, _, _ in audit['frete_divergente']: pedidos_problema.add(n)
n_ok = len(cvs) - len(pedidos_problema)

# === Workbook ===
wb = openpyxl.Workbook()
thin = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
              top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
red = PatternFill('solid', fgColor='FFE5DC')
yellow = PatternFill('solid', fgColor='FFF8E1')
green = PatternFill('solid', fgColor='E0F7F0')
blue = PatternFill('solid', fgColor='E8EAF6')

# === Aba 1: Resumo ===
ws = wb.active
ws.title = 'Resumo'
ws['A1'] = 'Auditoria Mi&Co - 2a quinzena abril/2026'
ws['A1'].font = Font(size=16, bold=True, color='2D3436')
ws.merge_cells('A1:C1')

ws['A3'] = 'Marca:'; ws['B3'] = 'Mi&Co (domain 1942933)'
ws['A4'] = 'Periodo:'; ws['B4'] = f'{de_str} a {ate_str}'
ws['A5'] = 'Total de pedidos:'; ws['B5'] = len(cvs)
ws['A6'] = 'Arquivo:'; ws['B6'] = 'MI E CO Corrigido.xlsx'

ws['A8'] = 'Resumo Financeiro'
ws['A8'].font = Font(size=13, bold=True, color='6C5CE7')

linhas_fin = [
    ('Total ValorPostagem (a pagar Onlog)', total_postagem, 'E17055', '<-- Vesti deve pagar a Onlog'),
    ('BIA total (cliente pagou no checkout)', total_bia, '0984E3', 'controle/referencia'),
    ('Minimo Postagem x 1.10', total_minimo, 'F39C12', 'piso de cobranca'),
    ('A COBRAR Mi&Co = MAX(BIA, P x 1.10)', total_cobrar, '00B894', '<-- valor a faturar para a marca'),
    ('Margem Vesti (a cobrar - postagem)', margem_vesti, '6C5CE7', 'lucro Vesti na quinzena'),
]
for i, (lbl, v, color, note) in enumerate(linhas_fin, start=9):
    ws.cell(row=i, column=1, value=lbl).font = Font(bold=True)
    cell = ws.cell(row=i, column=2, value=v)
    cell.number_format = '"R$ "#,##0.00'
    cell.font = Font(bold=True, color=color, size=12)
    ws.cell(row=i, column=3, value=note).font = Font(italic=True, color='636E72', size=11)

ws['A16'] = 'Auditoria - pontos de atencao'
ws['A16'].font = Font(size=13, bold=True, color='6C5CE7')

linhas_aud = [
    ('Pedidos OK (sem nenhum problema):', n_ok, '00B894'),
    ('Status atrasado no Mongo:', n_status_atr, 'F39C12'),
    ('Sem etiqueta no Mongo:', n_sem_etiq, 'E17055'),
    ('Frete divergente >5%:', n_frete_div, '0984E3'),
]
for i, (lbl, v, color) in enumerate(linhas_aud, start=17):
    ws.cell(row=i, column=1, value=lbl).font = Font(bold=True)
    cell = ws.cell(row=i, column=2, value=v)
    cell.font = Font(bold=True, color=color, size=12)

ws.column_dimensions['A'].width = 44
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 36

# === Aba 2: Detalhamento todos os pedidos ===
ws2 = wb.create_sheet('Pedidos')
cols2 = ['Pedido', 'Data', 'Operador', 'Modalidade', 'CodigoInterno', 'NF',
         'Cliente', 'Cidade', 'UF',
         'Status Onlog (planilha)', 'Status Mongo (Vesti)',
         'Etiqueta Mongo',
         'BIA (Mongo)', 'ValorPostagem (Onlog)', 'Postagem x 1.10',
         'A COBRAR (MAX)', 'Margem Vesti', 'Origem cobranca', 'Auditoria']
for ci, h in enumerate(cols2, 1):
    cell = ws2.cell(row=1, column=ci, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='6C5CE7')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for ri, (cv, c) in enumerate(sorted(cvs.items(), key=lambda x: int(x[0].split('_')[1])), start=2):
    n = cv.split('_')[1]
    minimo = c['postagem'] * 1.10
    cobrar = max(c['mongoBia'], minimo)
    margem = cobrar - c['postagem']
    origem = 'BIA (cliente pagou)' if c['mongoBia'] >= minimo - 0.01 else 'PISO 10%'

    flags = []
    if (c['mongoStatus'] or '').upper() in ('WAITING', 'SEPARATED') and c['statusOnlog']:
        flags.append('Status atrasado')
    if not c['mongoTemEtiqueta']:
        flags.append('Sem etiqueta Mongo')
    esperado = c['postagem'] * 1.10
    if c['mongoBia'] > 0 and c['postagem'] > 0:
        diff_pct = abs(c['mongoBia'] - esperado) / esperado
        if diff_pct > THRESHOLD_FRETE and abs(c['mongoBia'] - esperado) > 1.0:
            flags.append(f'Frete {diff_pct*100:.1f}%')
    audit_str = ' | '.join(flags) if flags else 'OK'
    fill = green if not flags else (red if 'Sem etiqueta' in audit_str else yellow)

    data_str = c['data'].strftime('%d/%m/%Y') if hasattr(c['data'], 'strftime') else str(c['data'])[:10]
    operador = c['provider'] or c['op']
    etiq_str = 'SIM' if c['mongoTemEtiqueta'] else 'NAO'
    vals = [n, data_str, operador, c['modal'], c['codigoInterno'], str(c['numeroNF'] or ''),
            c['cliente'], c['cidade'], c['uf'],
            c['statusOnlog'], c['mongoStatus'], etiq_str,
            c['mongoBia'], c['postagem'], minimo,
            cobrar, margem, origem, audit_str]
    for ci, v in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell.fill = fill
        cell.border = thin
        if ci in (13, 14, 15, 16, 17):
            cell.number_format = '"R$ "#,##0.00'
        if ci == 12:
            cell.font = Font(bold=True, color='00B894' if etiq_str == 'SIM' else 'E17055')
        if ci == 16:
            cell.font = Font(bold=True, color='00B894')
        if ci == 17:
            cell.font = Font(bold=True, color='6C5CE7')
        if ci == 18:
            color = '00B894' if origem.startswith('BIA') else 'E17055'
            cell.font = Font(bold=True, color=color)
        if ci == 19:
            color = '00B894' if audit_str == 'OK' else 'E17055'
            cell.font = Font(bold=True, color=color)

# Total
trow = len(cvs) + 2
ws2.cell(row=trow, column=1, value='TOTAL').font = Font(bold=True, size=12)
totals = {13: total_bia, 14: total_postagem, 15: total_minimo, 16: total_cobrar, 17: margem_vesti}
for ci, v in totals.items():
    cell = ws2.cell(row=trow, column=ci, value=v)
    cell.font = Font(bold=True)
    cell.number_format = '"R$ "#,##0.00'
    cell.fill = PatternFill('solid', fgColor='F0EFFF')
ws2.cell(row=trow, column=16).font = Font(bold=True, color='00B894', size=12)
ws2.cell(row=trow, column=17).font = Font(bold=True, color='6C5CE7', size=12)
ws2.cell(row=trow, column=1).fill = PatternFill('solid', fgColor='F0EFFF')

widths = [8, 11, 14, 12, 16, 14, 28, 18, 5, 28, 16, 9, 13, 14, 14, 14, 14, 18, 28]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 32
ws2.freeze_panes = 'A2'

# === Aba 3: Auditoria - Status atrasado ===
ws3 = wb.create_sheet('1. Status atrasado')
ws3['A1'] = '1. Status atrasado: Mongo nao acompanha Onlog (operacao precisa atualizar)'
ws3['A1'].font = Font(bold=True, size=12, color='F39C12')
ws3.merge_cells('A1:E1')
heads3 = ['Pedido', 'Cliente', 'Status Mongo (Vesti)', 'Status Onlog (planilha)', 'Operador']
for ci, h in enumerate(heads3, 1):
    cell = ws3.cell(row=3, column=ci, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='F39C12')
    cell.alignment = Alignment(horizontal='center')
for ri, (n, c) in enumerate(audit['status_atrasado'], start=4):
    vals = [n, c['cliente'], c['mongoStatus'], c['statusOnlog'], c['provider'] or c['op']]
    for ci, v in enumerate(vals, 1):
        cell = ws3.cell(row=ri, column=ci, value=v)
        cell.border = thin
        cell.fill = yellow
for i, w in enumerate([8, 28, 22, 30, 16], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# === Aba 4: Sem etiqueta ===
ws4 = wb.create_sheet('2. Sem etiqueta')
ws4['A1'] = '2. Sem etiqueta no Mongo: pedidos postados pela Onlog mas Vesti nao gravou URL da etiqueta'
ws4['A1'].font = Font(bold=True, size=12, color='E17055')
ws4.merge_cells('A1:F1')
heads4 = ['Pedido', 'Data', 'Cliente', 'Status Mongo', 'Status Onlog', 'Postagem']
for ci, h in enumerate(heads4, 1):
    cell = ws4.cell(row=3, column=ci, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='E17055')
    cell.alignment = Alignment(horizontal='center')
for ri, (n, c) in enumerate(audit['sem_etiqueta'], start=4):
    data_str = c['data'].strftime('%d/%m/%Y') if hasattr(c['data'], 'strftime') else ''
    vals = [n, data_str, c['cliente'], c['mongoStatus'], c['statusOnlog'], c['postagem']]
    for ci, v in enumerate(vals, 1):
        cell = ws4.cell(row=ri, column=ci, value=v)
        cell.border = thin
        cell.fill = red
        if ci == 6:
            cell.number_format = '"R$ "#,##0.00'
for i, w in enumerate([8, 11, 28, 18, 30, 12], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# === Aba 5: Frete divergente ===
ws5 = wb.create_sheet('3. Frete divergente')
ws5['A1'] = '3. Frete divergente >5%: BIA cobrado difere de Postagem x 1.10'
ws5['A1'].font = Font(bold=True, size=12, color='0984E3')
ws5.merge_cells('A1:G1')
heads5 = ['Pedido', 'Cliente', 'Postagem', 'Esperado (P x 1.10)', 'BIA cobrado', 'Diferenca', '%']
for ci, h in enumerate(heads5, 1):
    cell = ws5.cell(row=3, column=ci, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='0984E3')
    cell.alignment = Alignment(horizontal='center')
audit['frete_divergente'].sort(key=lambda x: -abs(x[2]))
for ri, (n, c, diff, pct) in enumerate(audit['frete_divergente'], start=4):
    esperado = c['postagem'] * 1.10
    color = '00B894' if diff >= 0 else 'E17055'
    vals = [n, c['cliente'], c['postagem'], esperado, c['mongoBia'], diff, pct]
    for ci, v in enumerate(vals, 1):
        cell = ws5.cell(row=ri, column=ci, value=v)
        cell.border = thin
        cell.fill = blue
        if ci in (3, 4, 5, 6):
            cell.number_format = '"R$ "#,##0.00'
        if ci == 7:
            cell.number_format = '0.0%'
        if ci in (6, 7):
            cell.font = Font(bold=True, color=color)
for i, w in enumerate([8, 28, 12, 16, 14, 14, 8], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# Save
out = rf'C:/Users/Laura/Downloads/Auditoria_MiCo_2quinzena_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx'
wb.save(out)
print(f'Salvo em: {out}')
print()
print(f'Pedidos: {len(cvs)}')
print(f'  OK:                   {n_ok}')
print(f'  Status atrasado:      {n_status_atr}')
print(f'  Sem etiqueta Mongo:   {n_sem_etiq}')
print(f'  Frete divergente >5%: {n_frete_div}')
print()
print(f'A pagar Onlog:    R$ {total_postagem:,.2f}')
print(f'A cobrar Mi&Co:   R$ {total_cobrar:,.2f}')
print(f'Margem Vesti:     R$ {margem_vesti:,.2f}')

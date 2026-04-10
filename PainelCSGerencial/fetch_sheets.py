"""
Extrai dados das planilhas do Google Sheets (baixadas como xlsx live):
- Pesquisa NPS / aba 'Nota NPS'           -> NPS_DATA
- OKRs e KPIs - Tribo 1 / aba 'Oraculo'   -> CSAT_ORACULO_DATA (row 22)

Baixa direto via 'export?format=xlsx' (precisa que o sheet esteja "Qualquer
pessoa com o link pode visualizar"). Salva os xlsx localmente e depois parseia.
"""

import json
import re
import sys
import io
import urllib.request
import urllib.error
from datetime import datetime, date
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).parent
NPS_XLSX = ROOT / "Pesquisa NPS.xlsx"
OKR_XLSX = ROOT / "OKRs e KPIs - Tribo 1.xlsx"
OUT = ROOT / "sheets_data.json"
SA_JSON = ROOT / "google_sa.json"

# Google Sheets IDs
NPS_SHEET_ID = "1TEA7_hVMFF3CynXn-d9uBXOAckcjueT8-cEk9g5Zzm0"
OKR_SHEET_ID = "15i3jOD_I6oAgtKUo34rliVZ3cqIQTPE9eM5Qm9sopWY"

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


def get_sa_token() -> str:
    if not SA_JSON.exists():
        raise RuntimeError(
            f"{SA_JSON.name} nao encontrado. Baixe a chave JSON da service "
            "account e salve nesse caminho. Tambem compartilhe os 2 sheets "
            "com o email da SA (Viewer)."
        )
    from google.oauth2 import service_account
    import google.auth.transport.requests
    creds = service_account.Credentials.from_service_account_file(
        str(SA_JSON), scopes=SCOPES
    )
    creds.refresh(google.auth.transport.requests.Request())
    return creds.token


def download_xlsx(sheet_id: str, dest: Path, token: str) -> None:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    print(f"[download] {dest.name} <- gsheet {sheet_id[:10]}...")
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {token}",
        "User-Agent": "Mozilla/5.0",
    })
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = resp.read()
    except urllib.error.HTTPError as e:
        msg = e.read().decode("utf-8", errors="replace")[:300]
        if e.code == 404:
            raise RuntimeError(
                f"HTTP 404: a service account nao tem acesso ao sheet {sheet_id}. "
                f"Compartilhe o sheet com o email da SA (Viewer)."
            )
        raise RuntimeError(f"HTTP {e.code} ao baixar {dest.name}: {msg}")
    if data[:2] != b"PK":
        snippet = data[:200].decode("utf-8", errors="replace")
        raise RuntimeError(f"Resposta nao eh xlsx valido: {snippet}")
    dest.write_bytes(data)
    print(f"[download] {dest.name} OK ({len(data)//1024} KB)")


def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v) -> int | None:
    f = _float(v)
    return int(f) if f is not None else None


def _normalize_phone(raw: str) -> str:
    """So digitos. Remove prefixo 55 (BR) se presente."""
    digits = re.sub(r"\D", "", str(raw or ""))
    if len(digits) > 11 and digits.startswith("55"):
        digits = digits[2:]
    return digits


def extract_nps() -> list[dict]:
    print(f"[nps] lendo {NPS_XLSX.name}")
    wb = openpyxl.load_workbook(NPS_XLSX, data_only=True)
    ws = wb["Nota NPS"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col(name: str) -> int | None:
        for i, h in enumerate(headers, 1):
            if h and h.strip() == name:
                return i
        return None

    c_nome = col("Nome")
    c_tel = col("Telefone")
    c_nota = col("Nota")
    c_data_pree = col("Data Preenchido")
    c_empresa = col("Nome Empresa")
    c_coment = col("Comentario")
    c_anjo = col("CS - Anjo")
    c_dom = col("Dominio")

    out = []
    for r in range(2, ws.max_row + 1):
        nota = _int(ws.cell(r, c_nota).value) if c_nota else None
        if nota is None:
            continue
        data_pree = ws.cell(r, c_data_pree).value if c_data_pree else None
        if isinstance(data_pree, str):
            data_iso = data_pree[:10]
        elif isinstance(data_pree, (datetime, date)):
            data_iso = data_pree.strftime("%Y-%m-%d")
        else:
            data_iso = ""

        phone = _s(ws.cell(r, c_tel).value) if c_tel else ""
        out.append({
            "nome": _s(ws.cell(r, c_nome).value) if c_nome else "",
            "telefone": phone,
            "telefone_norm": _normalize_phone(phone),
            "nota": nota,
            "data": data_iso,
            "empresa": _s(ws.cell(r, c_empresa).value) if c_empresa else "",
            "comentario": _s(ws.cell(r, c_coment).value) if c_coment else "",
            "anjo": _s(ws.cell(r, c_anjo).value) if c_anjo else "",
            "dominio": _s(ws.cell(r, c_dom).value) if c_dom else "",
        })
    print(f"[nps] {len(out)} respostas")
    return out


def _col_date(ws, col_idx: int) -> str:
    """Infere a data de uma coluna do Oraculo.
    Row 1 = ano (Year), row 2 = mes (Month), row 3 = dia numerico.
    Busca para a esquerda ate encontrar valores preenchidos (celulas mescladas).
    """
    year = None
    month = None
    day = None
    for c in range(col_idx, 0, -1):
        if year is None and ws.cell(1, c).value is not None:
            year = _int(ws.cell(1, c).value)
        if month is None and ws.cell(2, c).value is not None:
            m = ws.cell(2, c).value
            if isinstance(m, str):
                month = m.strip()
        if year is not None and month is not None:
            break
    day = _int(ws.cell(3, col_idx).value)
    return f"{year or '?'}-{month or '?'}" + (f"-{day}" if day else "")


MONTHS_PT = {
    "Janeiro": "01", "Fevereiro": "02", "Março": "03", "Marco": "03",
    "Abril": "04", "Maio": "05", "Junho": "06", "Julho": "07",
    "Agosto": "08", "Setembro": "09", "Outubro": "10",
    "Novembro": "11", "Dezembro": "12",
}


def _col_yyyymm(ws, col_idx: int) -> str:
    year = None
    month = None
    for c in range(col_idx, 0, -1):
        if year is None and ws.cell(1, c).value is not None:
            year = _int(ws.cell(1, c).value)
        if month is None:
            m = ws.cell(2, c).value
            if isinstance(m, str) and m.strip() in MONTHS_PT:
                month = MONTHS_PT[m.strip()]
        if year and month:
            break
    if year and month:
        return f"{year}-{month}"
    return ""


def extract_csat_oraculo() -> list[dict]:
    print(f"[csat-oraculo] lendo {OKR_XLSX.name}")
    wb = openpyxl.load_workbook(OKR_XLSX, data_only=True)
    ws = wb["Oráculo"]

    # Localiza a row do CSAT (col A == 'CSAT')
    csat_row = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip().lower() == "csat":
            csat_row = r
            break
    if not csat_row:
        raise RuntimeError("Linha 'CSAT' nao encontrada na aba Oraculo")
    print(f"[csat-oraculo] linha CSAT = row {csat_row}")

    out = []
    skipped = 0
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(csat_row, c)
        if not cell.comment:
            continue
        yyyymm = _col_yyyymm(ws, c)
        week_avg = _float(cell.value) if isinstance(cell.value, (int, float)) else None

        # Parse comment: linhas no formato "NOTA - EMPRESA [opcional coment]"
        txt = cell.comment.text
        for line in txt.split("\n"):
            line = line.strip()
            if not line:
                continue
            m = re.match(r"^(\d+(?:[.,]\d+)?)\s*[-–—]\s*(.+)$", line)
            if not m:
                skipped += 1
                continue
            try:
                nota = float(m.group(1).replace(",", "."))
            except ValueError:
                skipped += 1
                continue
            empresa_bruta = m.group(2).strip()
            # Se houver parenteses com observacao, separar
            obs = ""
            m2 = re.match(r"^([^()]+?)\s*\((.+)\)\s*$", empresa_bruta)
            if m2:
                empresa = m2.group(1).strip()
                obs = m2.group(2).strip()
            else:
                empresa = empresa_bruta
            out.append({
                "empresa": empresa,
                "nota": nota,
                "mes": yyyymm,
                "observacao": obs,
                "week_avg": week_avg,
                "col": cell.coordinate,
            })
    print(f"[csat-oraculo] {len(out)} notas parseadas ({skipped} linhas descartadas)")
    return out


def main() -> None:
    token = get_sa_token()
    download_xlsx(NPS_SHEET_ID, NPS_XLSX, token)
    download_xlsx(OKR_SHEET_ID, OKR_XLSX, token)
    nps = extract_nps()
    csat_oraculo = extract_csat_oraculo()
    data = {
        "nps": nps,
        "csat_oraculo": csat_oraculo,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[write] {OUT.name}")


if __name__ == "__main__":
    main()

"""
Ingere a planilha do Diogo (Fechamento ONLOG.xlsx) e gera onlog_diff.json,
que e' inlinado no dashboard como ONLOG_DIFF e mostra a conferencia da quinzena.

Workflow toda quinzena:
    1. Receber a planilha do Diogo
    2. py ingest_diogo_onlog.py "C:\\caminho\\Fechamento ONLOG.xlsx"
    3. py merge_data.py && py build_html.py
    4. git add -A && git commit -m "Onlog: conferencia <quinzena>" && git push

Saida:
    onlog_diff.json
        {
            "quinzena": {"de": "...", "ate": "..."},
            "geradoEm": "...",
            "planilhaArquivo": "Fechamento ONLOG.xlsx",
            "resumo": {"ok": N, "divergencias": N, "soPlanilha": N, "soFabric": N,
                       "nPlanilha": N, "nFabric": N},
            "divergencias": [{"orderNumber":..., "marca":..., "campo":..., "planilha":..., "fabric":...}, ...],
            "soPlanilha":   [{"codigoVolume":..., "orderNumber":..., "cliente":..., "cidade":..., "uf":..., "status":..., "postagem":...}, ...],
            "soFabric":     [{"data":..., "orderNumber":..., "marca":..., "cliente":..., "cidade":..., "uf":..., "cancelado":..., "status":..., "valor":...}, ...]
        }
"""

import argparse
import json
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl nao instalado. Rode: py -m pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).parent
ONLOG_JSON = ROOT / "onlog_data.json"
OUT_JSON = ROOT / "onlog_diff.json"


def norm_txt(s) -> str:
    if s is None:
        return ""
    nfd = unicodedata.normalize("NFD", str(s))
    no_diac = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    return " ".join("".join(c if c.isalnum() else " " for c in no_diac.upper()).split())


def norm_uf(s) -> str:
    return (str(s or "").strip().upper())[:2]


def parse_val_br(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def read_planilha(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        return []
    header = [str(h).strip() if h else "" for h in rows_raw[0]]
    out = []
    for r in rows_raw[1:]:
        out.append({header[i]: r[i] for i in range(len(header))})
    return out


def detect_quinzena(rows: list[dict]) -> tuple[str, str]:
    min_d = None
    for r in rows:
        d = r.get("Data")
        s = ""
        if isinstance(d, datetime):
            s = d.date().isoformat()
        elif isinstance(d, str):
            s = d[:10]
        if not s:
            continue
        if min_d is None or s < min_d:
            min_d = s
    if not min_d:
        return "", ""
    y, mo, dia = int(min_d[:4]), int(min_d[5:7]), int(min_d[8:10])
    mes = min_d[:7]
    if dia <= 15:
        return f"{mes}-01", f"{mes}-15"
    from calendar import monthrange
    return f"{mes}-16", f"{mes}-{monthrange(y, mo)[1]:02d}"


def detect_diogo_total(rows: list[dict]) -> float | None:
    """Detecta a linha totalizadora da planilha do Diogo:
    todos os campos vazios EXCETO ValorPostagem. Esse e' o valor que ele cobra."""
    for r in rows:
        if r.get("Destinatario") or r.get("CodigoInterno") or r.get("CodigoVolume"):
            continue
        if r.get("Remetente") or r.get("NumeroPedido") or r.get("NumeroNF"):
            continue
        v = parse_val_br(r.get("ValorPostagem"))
        if v is not None and v > 100:  # totalizador sempre tem valor alto
            return v
    return None


def aggregate_planilha(rows: list[dict], de: str, ate: str) -> tuple[dict, list]:
    """Retorna (pedidos_por_codigovolume, pa_vesti_avulsas).

    PA VESTI = linhas sem CodigoVolume (sem NumeroPedido) - postagens manuais
    geradas pela equipe Vesti direto no painel Onlog/Jadlog.
    """
    by = {}
    pa = []
    for r in rows:
        cv = str(r.get("CodigoVolume") or "").strip()
        d = r.get("Data")
        d_str = ""
        if isinstance(d, datetime):
            d_str = d.date().isoformat()
        elif isinstance(d, str):
            d_str = d[:10]
        if de and d_str and d_str < de:
            continue
        if ate and d_str and d_str > ate:
            continue
        v = parse_val_br(r.get("ValorPostagem"))
        if not cv or "_" not in cv:
            # PA VESTI - postagem avulsa sem pedido vinculado
            if v is not None and (r.get("Destinatario") or r.get("CodigoInterno")):
                pa.append({
                    "data": d_str,
                    "operador": r.get("Operador") or "",
                    "modalidade": r.get("Modalidade") or "",
                    "codigoInterno": r.get("CodigoInterno") or "",
                    "numeroNF": str(r.get("NumeroNF") or ""),
                    "remetente": r.get("Remetente") or "",
                    "cliente": r.get("Destinatario") or "",
                    "cidade": r.get("CidadeDestinatario") or "",
                    "uf": r.get("UFDestinatario") or "",
                    "status": r.get("Status") or "",
                    "postagem": round(v, 2),
                })
            continue
        if cv not in by:
            dom, order = cv.split("_", 1)
            by[cv] = {
                "codigoVolume": cv,
                "orderNumber": r.get("NumeroPedido") or order,
                "domainId": dom,
                "cliente": r.get("Destinatario") or "",
                "cidade": r.get("CidadeDestinatario") or "",
                "uf": r.get("UFDestinatario") or "",
                "status": r.get("Status") or "",
                "data": d_str,
                "postagem": 0.0,
            }
        if v is not None:
            by[cv]["postagem"] += v
    return by, pa


def filter_fabric(pedidos: list[dict], de: str, ate: str) -> dict:
    out = {}
    for p in pedidos:
        d = p.get("data") or ""
        if not d or d < de or d > ate:
            continue
        out[f'{p.get("dominioId","")}_{p.get("orderNumber","")}'] = p
    return out


def _is_no_postavel(p: dict) -> bool:
    """Pedidos cancelados ou ainda em SEPARATED nao deveriam aparecer na planilha
    do Diogo - sao 'so no Fabric' esperado e nao representam problema."""
    if p.get("cancelado"):
        return True
    return (p.get("status") or "").upper() == "SEPARATED"


def patch_onlog_data(onlog_data: dict, planilha: dict, de: str, ate: str) -> tuple[int, int]:
    """Atualiza valorPostagem e margemOnlog dos pedidos no onlog_data.json
    usando os valores reais da planilha do Diogo, para o range [de, ate].

    Margem = Cotacao BIA - Valor Postagem (lucro real da Vesti por frete).

    Retorna (n_atualizados, n_no_range_sem_planilha).
    """
    n_upd = 0
    n_skip = 0
    for p in onlog_data.get("pedidos", []):
        d = p.get("data") or ""
        if not d or d < de or d > ate:
            continue
        k = f'{p.get("dominioId","")}_{p.get("orderNumber","")}'
        pl = planilha.get(k)
        if not pl:
            n_skip += 1
            continue
        post = round(pl["postagem"], 2)
        p["valorPostagem"] = post
        p["postagemFonte"] = "planilha-diogo"
        # Status REAL da postagem (Diogo) - sobrepoe o status do Mongo no display
        if pl.get("status"):
            p["statusOnlog"] = pl["status"]
        bia = p.get("cotacaoBia")
        bia_f = float(bia) if bia is not None else None
        # Margem = Cotacao BIA - Valor Postagem (lucro real)
        if bia_f is not None and bia_f > 0:
            p["margemOnlog"] = round(bia_f - post, 2)
        else:
            # BIA zerado/ausente -> margem nao calculavel (frete gratis ou erro de cadastro)
            p["margemOnlog"] = None
        # Valor Ana FINAL = MAX(Cotacao BIA, Valor Postagem * 1.10)
        if bia_f is not None and bia_f > 0:
            p["valorAnaFinal"] = round(max(bia_f, post * 1.10), 2)
        else:
            p["valorAnaFinal"] = round(post * 1.10, 2)
        n_upd += 1
    return n_upd, n_skip


def fmt_brl(v) -> str:
    if v is None:
        return "-"
    return "R$ " + f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def compare(planilha: dict, fabric: dict) -> tuple[int, list, list, list]:
    keys = set(planilha) | set(fabric)
    dif, only_p, only_f, ok = [], [], [], 0
    for k in keys:
        pl = planilha.get(k)
        fa = fabric.get(k)
        if pl and not fa:
            only_p.append(pl)
            continue
        if fa and not pl:
            # cancelados e SEPARATED nao deveriam estar na planilha por design - ignora
            if _is_no_postavel(fa):
                continue
            only_f.append(fa)
            continue
        # Ignoramos cliente/destino - sao diferencas de formatacao/abreviacao
        # que nao representam erro real. Comparamos so o que importa: cancelamento e valor postagem.
        divs = []
        if fa.get("cancelado"):
            divs.append(("Cancelado", "(postado pela Onlog)", "CANCELADO no Vesti"))
        pp = pl["postagem"]
        pf = fa.get("valorPostagem")
        if pp is not None and pf is not None:
            if abs(pp - pf) > 0.01:
                divs.append(("Valor Postagem", fmt_brl(pp), fmt_brl(pf)))
        elif pp is not None and pf is None:
            divs.append(("Valor Postagem", fmt_brl(pp), "(sem dado no Fabric)"))
        if divs:
            for campo, a, b in divs:
                dif.append({
                    "orderNumber": fa.get("orderNumber"),
                    "marca": fa.get("marca", "-"),
                    "campo": campo,
                    "planilha": a,
                    "fabric": b,
                })
        else:
            ok += 1
    return ok, dif, only_p, only_f


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Caminho da planilha do Diogo")
    ap.add_argument("--de", help="Data inicial (default: detectado)")
    ap.add_argument("--ate", help="Data final (default: detectado)")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERRO: nao achei {xlsx_path}")
        sys.exit(1)
    if not ONLOG_JSON.exists():
        print(f"ERRO: {ONLOG_JSON} nao existe. Rode py fetch_onlog.py antes.")
        sys.exit(1)

    print(f"[1/4] Lendo planilha: {xlsx_path.name}")
    raw = read_planilha(xlsx_path)
    print(f"      {len(raw)} linhas brutas")
    diogo_total = detect_diogo_total(raw)
    if diogo_total is not None:
        print(f"      [linha totalizadora detectada] Diogo cobra: R$ {diogo_total:,.2f}")

    if args.de and args.ate:
        de, ate = args.de, args.ate
    else:
        de, ate = detect_quinzena(raw)
        if not de:
            print("ERRO: nao consegui detectar quinzena. Use --de/--ate.")
            sys.exit(1)
    print(f"      quinzena: {de} a {ate}")

    print(f"[2/4] Agregando planilha (CodigoVolume)")
    planilha, pa_vesti = aggregate_planilha(raw, de, ate)
    pa_total = round(sum(p["postagem"] for p in pa_vesti), 2)
    print(f"      {len(planilha)} pedidos unicos com CodigoVolume")
    print(f"      {len(pa_vesti)} postagens avulsas PA VESTI (sem pedido) - total R$ {pa_total:,.2f}")

    print(f"[3/4] Lendo Fabric (onlog_data.json)")
    onlog_data = json.loads(ONLOG_JSON.read_text(encoding="utf-8"))
    fabric = filter_fabric(onlog_data.get("pedidos", []), de, ate)
    print(f"      {len(fabric)} pedidos do Fabric na quinzena")

    print(f"[3.5/4] Patch onlog_data.json com valores da planilha (postagem + margem)")
    n_upd, n_skip = patch_onlog_data(onlog_data, planilha, de, ate)
    ONLOG_JSON.write_text(json.dumps(onlog_data, ensure_ascii=False), encoding="utf-8")
    print(f"      {n_upd} pedidos atualizados (postagem + margem); {n_skip} sem match na planilha")

    print(f"[4/4] Comparando")
    ok, dif, only_p, only_f = compare(planilha, fabric)
    n_dif_uniq = len({d["orderNumber"] for d in dif})
    print(f"      OK={ok}  Divergencias={len(dif)} ({n_dif_uniq} pedidos)")
    print(f"      So planilha={len(only_p)}  So Fabric={len(only_f)}")

    # totais financeiros
    total_pedidos_planilha = round(sum(p["postagem"] for p in planilha.values()), 2)
    # cobranca por marca (postagem * 1.10) - so para pedidos com domainId valido
    cobranca_marca = {}
    for p in planilha.values():
        dom = p["domainId"]
        cobranca_marca.setdefault(dom, {"domainId": dom, "nPedidos": 0, "postagem": 0.0, "cobrar": 0.0})
        cobranca_marca[dom]["nPedidos"] += 1
        cobranca_marca[dom]["postagem"] += p["postagem"]
        cobranca_marca[dom]["cobrar"] += p["postagem"] * 1.10
    # Anexar nome da marca via companies_data.json
    companies_path = ROOT / "companies_data.json"
    if companies_path.exists():
        cs = json.loads(companies_path.read_text(encoding="utf-8"))
        nome_por_dom = {}
        for c in cs:
            d = str(c.get("domain_id") or "")
            if not d:
                continue
            if c.get("isMatriz") or d not in nome_por_dom:
                nome_por_dom[d] = c.get("nome_fantasia") or c.get("name") or ""
        for dom, info in cobranca_marca.items():
            info["marca"] = nome_por_dom.get(dom, "")
    cobranca_lista = sorted(
        [{"domainId": d, "marca": v.get("marca", ""), "nPedidos": v["nPedidos"],
          "postagem": round(v["postagem"], 2), "cobrar": round(v["cobrar"], 2)}
         for d, v in cobranca_marca.items()],
        key=lambda x: -x["cobrar"]
    )

    total_geral = round(total_pedidos_planilha + pa_total, 2)
    # Custo real = o que o Diogo cobra (linha totalizadora). Fallback: soma da planilha.
    custo_diogo = round(diogo_total, 2) if diogo_total is not None else total_geral
    # Receita = custo Diogo * 1.10 (Vesti cobra das marcas o que paga + 10%)
    total_cobrar = round(custo_diogo * 1.10, 2)
    out = {
        "quinzena": {"de": de, "ate": ate},
        "geradoEm": datetime.now().isoformat(),
        "planilhaArquivo": xlsx_path.name,
        "resumo": {
            "ok": ok,
            "divergencias": len(dif),
            "divergenciasPedidos": n_dif_uniq,
            "soPlanilha": len(only_p),
            "soFabric": len(only_f),
            "nPlanilha": len(planilha),
            "nFabric": len(fabric),
            "totalPedidosPostagem": total_pedidos_planilha,
            "totalPaVesti": pa_total,
            "totalGeralPostagem": total_geral,
            "totalCobrarPedidos": round(total_pedidos_planilha * 1.10, 2),
            "totalCobrarPa": round(pa_total * 1.10, 2),
            "totalCobrarMarcas": total_cobrar,
            "custoDiogo": custo_diogo,
            "diogoDetectado": diogo_total is not None,
            "margemVesti": round(total_cobrar - custo_diogo, 2),
            "nPaVesti": len(pa_vesti),
        },
        "paVesti": pa_vesti,
        "cobrancaPorMarca": cobranca_lista,
        "divergencias": dif,
        "soPlanilha": [{
            "codigoVolume": p["codigoVolume"],
            "orderNumber": p["orderNumber"],
            "cliente": p["cliente"],
            "cidade": p["cidade"],
            "uf": p["uf"],
            "status": p["status"],
            "postagem": round(p["postagem"], 2),
        } for p in only_p],
        "soFabric": [{
            "data": p.get("data"),
            "orderNumber": p.get("orderNumber"),
            "marca": p.get("marca", "-"),
            "cliente": p.get("cliente", "-"),
            "cidade": p.get("cidade", ""),
            "uf": p.get("uf", ""),
            "cancelado": bool(p.get("cancelado")),
            "status": p.get("status", ""),
            "valor": p.get("valor"),
        } for p in only_f],
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n>> {OUT_JSON.name} escrito ({OUT_JSON.stat().st_size//1024} KB)")
    print(f">> Agora rode: py merge_data.py && py build_html.py")


if __name__ == "__main__":
    main()
